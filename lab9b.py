import openpyxl

wb=openpyxl.Workbook()
sheet=wb.active
sheet.title="Language"
lang=['Kannnada','Telugu','Tamil','Malayalam','Konkani','Marathi']
capital=['Bengaluru','Hyderabad','Chennai','Thiruvanthapuram','Panaji','Mmumbai']
code=['KA','TS','TN','KL','GA','MH']
state=['Karnataka','Telangana','Tamilnadu','Kerala','Goa','Maharashtra']
sheet.cell(1,1,value="State")
sheet.cell(1,2,value="Language")
sheet.cell(1,3,value="Capital")
sheet.cell(1,4,value="Code")
for i in range(2,8):
    sheet.cell(i,1,value=state[i-2])
    sheet.cell(i,2,value=lang[i-2])
    sheet.cell(i,3,value=capital[i-2])
    sheet.cell(i,4,value=code[i-2])
wb.save("demo.xlsx")
print("xlsheet created successfully")
print("Reading from XL sheet")
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
    print("\n")
    for cell in row:
        print(cell.value,end=' ')
